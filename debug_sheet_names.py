#!/usr/bin/env python3
"""
Debug script to check Excel sheet names
"""

import openpyxl
import os

def check_excel_sheets():
    excel_file = './data/workholic-data.xlsx'
    
    if not os.path.exists(excel_file):
        print("❌ Excel file does not exist")
        return
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        print(f"✅ Excel file loaded successfully")
        print(f"Sheet names: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\nSheet '{sheet_name}':")
            print(f"  Max row: {ws.max_row}")
            print(f"  Max column: {ws.max_column}")
            
            # Show first few rows
            for row in range(1, min(4, ws.max_row + 1)):
                row_data = []
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value else '')
                print(f"  Row {row}: {row_data}")
                
    except Exception as e:
        print(f"❌ Error: {e}")

if __name__ == "__main__":
    check_excel_sheets()
