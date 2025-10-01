from flask import Flask, request, jsonify, session, send_file, render_template
from flask_cors import CORS
import openpyxl
from openpyxl import Workbook
import os
import json
from datetime import datetime, timedelta
import io

app = Flask(__name__)
app.secret_key = 'workholic-secret-key'  # Change this in production
CORS(app)

# Excel file path
EXCEL_FILE = './data/workholic-data.xlsx'

def read_excel_data():
    """Read data from Excel file"""
    try:
        if not os.path.exists(EXCEL_FILE):
            return {'employees': [], 'attendance': [], 'leaderboard': []}
        
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        
        # Read employees sheet
        employees = []
        if 'Employees' in workbook.sheetnames:
            employees_sheet = workbook['Employees']
            for row in employees_sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:  # If email exists
                    employees.append({
                        'email': row[0],
                        'name': row[1] or '',
                        'role': row[2],
                        'schedule': row[3],
                        'password': row[4] or None
                    })
        
        # Read attendance sheet
        attendance = []
        if 'Attendance' in workbook.sheetnames:
            attendance_sheet = workbook['Attendance']
            for row in attendance_sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:  # If email exists
                    breaks_data = row[4] if row[4] else '[]'
                    try:
                        breaks = json.loads(breaks_data) if breaks_data else []
                    except:
                        breaks = []
                    
                    attendance.append({
                        'email': row[0],
                        'date': row[1],
                        'clockIn': row[2],
                        'clockOut': row[3],
                        'breaks': breaks,
                        'status': row[5] or 'A'
                    })
        
        # Read leaderboard sheet
        leaderboard = []
        if 'Leaderboard' in workbook.sheetnames:
            leaderboard_sheet = workbook['Leaderboard']
            for row in leaderboard_sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:  # If email exists
                    leaderboard.append({
                        'email': row[0],
                        'name': row[1] or '',
                        'totalPoints': row[2] or 0,
                        'attendancePoints': row[3] or 0,
                        'smallTasks': row[4] or 0,
                        'regularTasks': row[5] or 0,
                        'bigTasks': row[6] or 0
                    })
        
        return {'employees': employees, 'attendance': attendance, 'leaderboard': leaderboard}
    
    except Exception as e:
        print(f'Error reading Excel file: {e}')
        return {'employees': [], 'attendance': [], 'leaderboard': []}

def write_excel_data(data):
    """Write data to Excel file"""
    try:
        # Create workbook
        workbook = Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Create Employees sheet
        employees_sheet = workbook.create_sheet('Employees')
        employees_sheet.append(['Email', 'Name', 'Role', 'Schedule', 'Password'])
        for employee in data['employees']:
            employees_sheet.append([
                employee['email'],
                employee['name'],
                employee['role'],
                employee['schedule'],
                employee['password'] or ''
            ])
        
        # Create Attendance sheet
        attendance_sheet = workbook.create_sheet('Attendance')
        attendance_sheet.append(['Email', 'Date', 'Clock In', 'Clock Out', 'Breaks', 'Status'])
        for record in data['attendance']:
            attendance_sheet.append([
                record['email'],
                record['date'],
                record['clockIn'] or '',
                record['clockOut'] or '',
                json.dumps(record['breaks']),
                record['status']
            ])
        
        # Create Leaderboard sheet
        leaderboard_sheet = workbook.create_sheet('Leaderboard')
        leaderboard_sheet.append(['Email', 'Name', 'Total Points', 'Attendance Points', 'Small Tasks', 'Regular Tasks', 'Big Tasks'])
        for entry in data['leaderboard']:
            leaderboard_sheet.append([
                entry['email'],
                entry['name'],
                entry['totalPoints'],
                entry['attendancePoints'],
                entry['smallTasks'],
                entry['regularTasks'],
                entry['bigTasks']
            ])
        
        # Style headers
        for sheet in [employees_sheet, attendance_sheet, leaderboard_sheet]:
            for cell in sheet[1]:
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        # Save file
        os.makedirs(os.path.dirname(EXCEL_FILE), exist_ok=True)
        workbook.save(EXCEL_FILE)
        return True
    
    except Exception as e:
        print(f'Error writing Excel file: {e}')
        return False

def initialize_excel_file():
    """Initialize Excel file if it doesn't exist"""
    if not os.path.exists(EXCEL_FILE):
        initial_data = {
            'employees': [
                {
                    'email': 'admin@workholic.in',
                    'password': 'admin123',
                    'name': 'Admin',
                    'role': 'admin',
                    'schedule': 'general'
                }
            ],
            'attendance': [],
            'leaderboard': []
        }
        write_excel_data(initial_data)
        print('Initial Excel file created')
    else:
        # Ensure admin user exists in existing file
        data = read_excel_data()
        admin_exists = False
        for user in data['employees']:
            if user['email'] == 'admin@workholic.in':
                admin_exists = True
                break
        
        if not admin_exists:
            admin_user = {
                'email': 'admin@workholic.in',
                'password': 'admin123',
                'name': 'Admin',
                'role': 'admin',
                'schedule': 'general'
            }
            data['employees'].append(admin_user)
            try:
                write_excel_data(data)
                print('Admin user added to existing database')
            except:
                print('Admin user added to memory (Excel file locked)')

def get_schedule_for_employee(schedule):
    """Get schedule configuration for employee"""
    schedules = {
        'general': {
            'weekdays': {'start': '10:30', 'end': '19:00', 'breakDuration': 60},
            'saturday': {'start': '10:00', 'end': '13:00', 'breakDuration': 15},
            'workingSaturdays': [1, 3]
        },
        'shreyas': {
            'weekdays': {'start': '16:30', 'end': '19:00', 'breakDuration': 15},
            'friday': {'start': '12:00', 'end': '18:00', 'breakDuration': 45},
            'weekend': 'off'
        },
        'srushti': {
            'weekdays': {'start': '10:30', 'end': '16:30', 'breakDuration': 45},
            'saturday': {'start': '10:00', 'end': '13:00', 'breakDuration': 15},
            'workingSaturdays': [1, 3]
        },
        'vinay': {
            'weekdays': {'start': '10:30', 'end': '21:00', 'breakDuration': 60},
            'saturday': {'start': '10:00', 'end': '13:00', 'breakDuration': 15},
            'workingSaturdays': [1, 3]
        }
    }
    return schedules.get(schedule, schedules['general'])

def calculate_attendance_status(clock_in, clock_out, schedule, total_break_time):
    """Calculate attendance status"""
    if not clock_in:
        return 'A'  # Absent
    
    clock_in_time = datetime.fromisoformat(clock_in.replace('Z', '+00:00'))
    expected_start = clock_in_time.replace(hour=int(schedule['start'].split(':')[0]), 
                                         minute=int(schedule['start'].split(':')[1]), 
                                         second=0, microsecond=0)
    
    # Check if late by more than 15 minutes
    late_by = (clock_in_time - expected_start).total_seconds() / 60
    
    if not clock_out:
        return 'HD' if late_by > 15 else 'HD'  # Incomplete day
    
    clock_out_time = datetime.fromisoformat(clock_out.replace('Z', '+00:00'))
    total_worked_minutes = (clock_out_time - clock_in_time).total_seconds() / 60 - (total_break_time or 0)
    
    # Calculate expected work hours
    expected_end = expected_start.replace(hour=int(schedule['end'].split(':')[0]), 
                                        minute=int(schedule['end'].split(':')[1]))
    expected_work_minutes = (expected_end - expected_start).total_seconds() / 60 - schedule['breakDuration']
    
    if late_by > 15:
        return 'HD'
    if total_worked_minutes < expected_work_minutes * 0.6:
        return 'HD'
    if total_worked_minutes < expected_work_minutes * 0.8:
        return 'HD'
    
    return 'FD'

def update_leaderboard_points(email, attendance_status, data):
    """Update leaderboard points for employee"""
    try:
        employee_entry = None
        for entry in data['leaderboard']:
            if entry['email'] == email:
                employee_entry = entry
                break
        
        if not employee_entry:
            # Find employee info
            employee = None
            for emp in data['employees']:
                if emp['email'] == email:
                    employee = emp
                    break
            
            employee_entry = {
                'email': email,
                'name': employee['name'] if employee else email,
                'totalPoints': 0,
                'attendancePoints': 0,
                'smallTasks': 0,
                'regularTasks': 0,
                'bigTasks': 0
            }
            data['leaderboard'].append(employee_entry)
        
        # Update attendance points
        points_map = {'FD': 2, 'HD': 1, 'A': -1}
        employee_entry['attendancePoints'] += points_map.get(attendance_status, 0)
        
        # Recalculate total points
        employee_entry['totalPoints'] = (
            employee_entry['attendancePoints'] +
            (employee_entry['smallTasks'] * 1) +
            (employee_entry['regularTasks'] * 2) +
            (employee_entry['bigTasks'] * 3)
        )
        
        write_excel_data(data)
        return True
    
    except Exception as e:
        print(f'Update leaderboard error: {e}')
        return False

# Routes
@app.route('/')
def index():
    """Main route - redirect based on user role"""
    if 'user' in session:
        if session['user']['role'] == 'admin':
            return render_template('admin.html')
        else:
            return render_template('employee.html')
    else:
        return render_template('login.html')

@app.route('/api/test')
def test():
    """Test endpoint"""
    return jsonify({'success': True, 'message': 'Server is running'})

@app.route('/api/login', methods=['POST'])
def login():
    """User login"""
    data = request.get_json()
    email = data.get('email')
    password = data.get('password')
    
    try:
        excel_data = read_excel_data()
        employee = None
        for emp in excel_data['employees']:
            if emp['email'] == email:
                employee = emp
                break
        
        if not employee:
            return jsonify({'success': False, 'message': 'Invalid credentials'})
        
        # First time login - set password
        if not employee['password']:
            employee['password'] = password
            excel_data['employees'] = [emp for emp in excel_data['employees'] if emp['email'] != email]
            excel_data['employees'].append(employee)
            write_excel_data(excel_data)
            
            session['user'] = {
                'email': employee['email'],
                'name': employee['name'],
                'role': employee['role'],
                'schedule': employee['schedule']
            }
            
            return jsonify({'success': True, 'firstLogin': True, 'role': employee['role']})
        
        # Regular login
        if employee['password'] != password:
            return jsonify({'success': False, 'message': 'Invalid credentials'})
        
        session['user'] = {
            'email': employee['email'],
            'name': employee['name'],
            'role': employee['role'],
            'schedule': employee['schedule']
        }
        
        return jsonify({'success': True, 'role': employee['role']})
    
    except Exception as e:
        print(f'Login error: {e}')
        return jsonify({'success': False, 'message': 'Server error'})

@app.route('/api/logout', methods=['POST'])
def logout():
    """User logout"""
    session.clear()
    return jsonify({'success': True})

@app.route('/api/user')
def get_user():
    """Get current user info"""
    if 'user' in session:
        return jsonify({'success': True, 'user': session['user']})
    else:
        return jsonify({'success': False})

@app.route('/api/clock-action', methods=['POST'])
def clock_action():
    """Handle clock in/out and break actions"""
    if 'user' not in session or session['user']['role'] != 'employee':
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    data = request.get_json()
    action = data.get('action')
    email = session['user']['email']
    now = datetime.now().isoformat()
    today = datetime.now().strftime('%Y-%m-%d')
    
    try:
        excel_data = read_excel_data()
        today_record = None
        
        for record in excel_data['attendance']:
            if record['email'] == email and record['date'] == today:
                today_record = record
                break
        
        if not today_record:
            today_record = {
                'email': email,
                'date': today,
                'clockIn': None,
                'clockOut': None,
                'breaks': [],
                'status': 'A'
            }
            excel_data['attendance'].append(today_record)
        
        if action == 'clock-in':
            if not today_record['clockIn']:
                today_record['clockIn'] = now
        
        elif action == 'break-start':
            last_break = today_record['breaks'][-1] if today_record['breaks'] else None
            if not last_break or last_break.get('end'):
                today_record['breaks'].append({'start': now, 'end': None})
        
        elif action == 'break-end':
            if today_record['breaks']:
                current_break = today_record['breaks'][-1]
                if current_break and not current_break.get('end'):
                    current_break['end'] = now
        
        elif action == 'clock-out':
            if today_record['clockIn']:
                today_record['clockOut'] = now
                
                # End any ongoing break
                if today_record['breaks']:
                    ongoing_break = today_record['breaks'][-1]
                    if ongoing_break and not ongoing_break.get('end'):
                        ongoing_break['end'] = now
                
                # Calculate attendance status
                schedule = get_schedule_for_employee(session['user']['schedule'])
                total_break_time = 0
                for brk in today_record['breaks']:
                    if brk.get('start') and brk.get('end'):
                        start_time = datetime.fromisoformat(brk['start'].replace('Z', '+00:00'))
                        end_time = datetime.fromisoformat(brk['end'].replace('Z', '+00:00'))
                        total_break_time += (end_time - start_time).total_seconds() / 60
                
                today_record['status'] = calculate_attendance_status(
                    today_record['clockIn'],
                    today_record['clockOut'],
                    schedule.get('weekdays', schedule),
                    total_break_time
                )
                
                # Update leaderboard points
                update_leaderboard_points(email, today_record['status'], excel_data)
        
        write_excel_data(excel_data)
        return jsonify({'success': True, 'record': today_record})
    
    except Exception as e:
        print(f'Clock action error: {e}')
        return jsonify({'success': False, 'message': 'Server error'})

@app.route('/api/attendance/today')
def get_today_attendance():
    """Get today's attendance record"""
    if 'user' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    email = session['user']['email']
    today = datetime.now().strftime('%Y-%m-%d')
    
    try:
        excel_data = read_excel_data()
        today_record = None
        
        for record in excel_data['attendance']:
            if record['email'] == email and record['date'] == today:
                today_record = record
                break
        
        return jsonify({'success': True, 'record': today_record})
    
    except Exception as e:
        print(f'Today attendance error: {e}')
        return jsonify({'success': False, 'message': 'Server error'})

@app.route('/api/attendance/history')
def get_attendance_history():
    """Get attendance history"""
    if 'user' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    email = session['user']['email']
    
    try:
        excel_data = read_excel_data()
        user_attendance = [record for record in excel_data['attendance'] if record['email'] == email]
        user_attendance.sort(key=lambda x: x['date'], reverse=True)
        user_attendance = user_attendance[:30]  # Last 30 days
        
        return jsonify({'success': True, 'attendance': user_attendance})
    
    except Exception as e:
        print(f'Attendance history error: {e}')
        return jsonify({'success': False, 'message': 'Server error'})

@app.route('/api/schedule/today')
def get_today_schedule():
    """Get today's schedule"""
    if 'user' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    try:
        schedule = get_schedule_for_employee(session['user']['schedule'])
        return jsonify({'success': True, 'schedule': schedule})
    
    except Exception as e:
        print(f'Schedule error: {e}')
        return jsonify({'success': False, 'message': 'Server error'})

@app.route('/api/leaderboard')
def get_leaderboard():
    """Get leaderboard data"""
    try:
        excel_data = read_excel_data()
        leaderboard = sorted(excel_data['leaderboard'], key=lambda x: x['totalPoints'], reverse=True)
        
        for i, entry in enumerate(leaderboard):
            entry['rank'] = i + 1
        
        top8 = leaderboard[:8]
        user_rank = None
        
        if 'user' in session and session['user']['role'] == 'employee':
            for entry in leaderboard:
                if entry['email'] == session['user']['email']:
                    if entry['rank'] > 8:
                        user_rank = entry['rank']
                    break
        
        return jsonify({'success': True, 'leaderboard': top8, 'userRank': user_rank})
    
    except Exception as e:
        print(f'Leaderboard error: {e}')
        return jsonify({'success': False, 'message': 'Server error'})

@app.route('/api/admin/update-tasks', methods=['POST'])
def update_tasks():
    """Update employee tasks (admin only)"""
    if 'user' not in session or session['user']['role'] != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    data = request.get_json()
    email = data.get('email')
    task_type = data.get('taskType')
    count = int(data.get('count', 0))
    
    try:
        excel_data = read_excel_data()
        employee_entry = None
        
        for entry in excel_data['leaderboard']:
            if entry['email'] == email:
                employee_entry = entry
                break
        
        if not employee_entry:
            # Find employee info
            employee = None
            for emp in excel_data['employees']:
                if emp['email'] == email:
                    employee = emp
                    break
            
            employee_entry = {
                'email': email,
                'name': employee['name'] if employee else email,
                'totalPoints': 0,
                'smallTasks': 0,
                'regularTasks': 0,
                'bigTasks': 0,
                'attendancePoints': 0
            }
            excel_data['leaderboard'].append(employee_entry)
        
        # Update task count
        task_counts = {'small': 'smallTasks', 'regular': 'regularTasks', 'big': 'bigTasks'}
        if task_type in task_counts:
            old_count = employee_entry.get(task_counts[task_type], 0)
            new_count = old_count + count
            employee_entry[task_counts[task_type]] = max(0, new_count)
            
            # Recalculate total points
            employee_entry['totalPoints'] = (
                (employee_entry.get('attendancePoints', 0)) +
                (employee_entry.get('smallTasks', 0) * 1) +
                (employee_entry.get('regularTasks', 0) * 2) +
                (employee_entry.get('bigTasks', 0) * 3)
            )
        
        write_excel_data(excel_data)
        return jsonify({'success': True})
    
    except Exception as e:
        print(f'Update tasks error: {e}')
        return jsonify({'success': False, 'message': 'Server error'})

@app.route('/api/admin/employees')
def get_employees():
    """Get employee list (admin only)"""
    if 'user' not in session or session['user']['role'] != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    try:
        excel_data = read_excel_data()
        employees = [{'email': emp['email'], 'name': emp['name']} 
                    for emp in excel_data['employees'] if emp['role'] == 'employee']
        
        return jsonify({'success': True, 'employees': employees})
    
    except Exception as e:
        print(f'Get employees error: {e}')
        return jsonify({'success': False, 'message': 'Server error'})

@app.route('/api/admin/attendance')
def get_admin_attendance():
    """Get all attendance records (admin only)"""
    if 'user' not in session or session['user']['role'] != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    try:
        excel_data = read_excel_data()
        
        attendance_with_names = []
        for record in excel_data['attendance']:
            employee = None
            for emp in excel_data['employees']:
                if emp['email'] == record['email']:
                    employee = emp
                    break
            
            attendance_with_names.append({
                **record,
                'name': employee['name'] if employee else record['email']
            })
        
        return jsonify({'success': True, 'attendance': attendance_with_names})
    
    except Exception as e:
        print(f'Admin attendance error: {e}')
        return jsonify({'success': False, 'message': 'Server error'})

@app.route('/api/admin/users')
def get_users():
    """Get all users (admin only)"""
    if 'user' not in session or session['user']['role'] != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    try:
        excel_data = read_excel_data()
        users = [{'email': emp['email'], 'name': emp['name'], 'role': emp['role'], 'schedule': emp['schedule']} 
                for emp in excel_data['employees']]
        
        return jsonify({'success': True, 'users': users})
    
    except Exception as e:
        print(f'Get users error: {e}')
        return jsonify({'success': False, 'message': 'Server error'})

@app.route('/api/admin/users/<email>')
def get_user_by_email(email):
    """Get specific user (admin only)"""
    if 'user' not in session or session['user']['role'] != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    try:
        excel_data = read_excel_data()
        user = None
        
        for emp in excel_data['employees']:
            if emp['email'] == email:
                user = emp
                break
        
        if not user:
            return jsonify({'success': False, 'message': 'User not found'})
        
        return jsonify({'success': True, 'user': {
            'email': user['email'],
            'name': user['name'],
            'role': user['role'],
            'schedule': user['schedule']
        }})
    
    except Exception as e:
        print(f'Get user error: {e}')
        return jsonify({'success': False, 'message': 'Server error'})

@app.route('/api/admin/users/create', methods=['POST'])
def create_user():
    """Create new user (admin only)"""
    if 'user' not in session or session['user']['role'] != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    data = request.get_json()
    name = data.get('name')
    email = data.get('email')
    role = data.get('role')
    schedule = data.get('schedule')
    password = data.get('password')
    
    try:
        excel_data = read_excel_data()
        
        # Check if user already exists
        for emp in excel_data['employees']:
            if emp['email'] == email:
                return jsonify({'success': False, 'message': 'User already exists'})
        
        new_user = {
            'email': email,
            'name': name or '',
            'role': role,
            'schedule': schedule,
            'password': password or None
        }
        
        excel_data['employees'].append(new_user)
        write_excel_data(excel_data)
        
        return jsonify({'success': True})
    
    except Exception as e:
        print(f'Create user error: {e}')
        return jsonify({'success': False, 'message': 'Server error'})

@app.route('/api/admin/users/update', methods=['POST'])
def update_user():
    """Update user (admin only)"""
    if 'user' not in session or session['user']['role'] != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    data = request.get_json()
    name = data.get('name')
    email = data.get('email')
    role = data.get('role')
    schedule = data.get('schedule')
    password = data.get('password')
    
    try:
        excel_data = read_excel_data()
        
        for i, emp in enumerate(excel_data['employees']):
            if emp['email'] == email:
                excel_data['employees'][i]['name'] = name or ''
                excel_data['employees'][i]['role'] = role
                excel_data['employees'][i]['schedule'] = schedule
                
                if password:
                    excel_data['employees'][i]['password'] = password
                
                write_excel_data(excel_data)
                return jsonify({'success': True})
        
        return jsonify({'success': False, 'message': 'User not found'})
    
    except Exception as e:
        print(f'Update user error: {e}')
        return jsonify({'success': False, 'message': 'Server error'})

@app.route('/api/admin/users/delete', methods=['POST'])
def delete_user():
    """Delete user (admin only)"""
    if 'user' not in session or session['user']['role'] != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    data = request.get_json()
    email = data.get('email')
    
    try:
        excel_data = read_excel_data()
        original_length = len(excel_data['employees'])
        
        excel_data['employees'] = [emp for emp in excel_data['employees'] if emp['email'] != email]
        
        if len(excel_data['employees']) == original_length:
            return jsonify({'success': False, 'message': 'User not found'})
        
        write_excel_data(excel_data)
        return jsonify({'success': True})
    
    except Exception as e:
        print(f'Delete user error: {e}')
        return jsonify({'success': False, 'message': 'Server error'})

@app.route('/api/admin/export-excel')
def export_excel():
    """Export attendance data to Excel (admin only)"""
    if 'user' not in session or session['user']['role'] != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    try:
        excel_data = read_excel_data()
        
        # Create workbook
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Attendance Data'
        
        # Add headers
        headers = ['Employee Name', 'Email', 'Date', 'Clock In', 'Clock Out', 'Status', 'Break Duration (min)', 'Total Breaks']
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        # Add data
        row = 2
        for record in excel_data['attendance']:
            # Find employee name
            employee = None
            for emp in excel_data['employees']:
                if emp['email'] == record['email']:
                    employee = emp
                    break
            
            name = employee['name'] if employee else record['email']
            
            # Calculate break duration
            break_duration = 0
            total_breaks = len(record.get('breaks', []))
            for brk in record.get('breaks', []):
                if brk.get('start') and brk.get('end'):
                    start_time = datetime.fromisoformat(brk['start'].replace('Z', '+00:00'))
                    end_time = datetime.fromisoformat(brk['end'].replace('Z', '+00:00'))
                    break_duration += (end_time - start_time).total_seconds() / 60
            
            clock_in = record['clockIn'].replace('T', ' ').replace('Z', '') if record['clockIn'] else 'N/A'
            clock_out = record['clockOut'].replace('T', ' ').replace('Z', '') if record['clockOut'] else 'N/A'
            
            worksheet.cell(row=row, column=1, value=name)
            worksheet.cell(row=row, column=2, value=record['email'])
            worksheet.cell(row=row, column=3, value=record['date'])
            worksheet.cell(row=row, column=4, value=clock_in)
            worksheet.cell(row=row, column=5, value=clock_out)
            worksheet.cell(row=row, column=6, value=record['status'])
            worksheet.cell(row=row, column=7, value=int(break_duration))
            worksheet.cell(row=row, column=8, value=total_breaks)
            
            row += 1
        
        # Style header row
        for col in range(1, len(headers) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        # Save to bytes
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=f'attendance-export-{datetime.now().strftime("%Y-%m-%d")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    except Exception as e:
        print(f'Excel export error: {e}')
        return jsonify({'success': False, 'message': 'Server error'})

if __name__ == '__main__':
    # Initialize Excel file
    initialize_excel_file()
    
    # Run the app
    app.run(host='0.0.0.0', port=5000, debug=True)
