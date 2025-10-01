#!/usr/bin/env python3
"""
Debug the login flow to see what's happening
"""

import os
import openpyxl
from openpyxl import Workbook

def debug_login_flow():
    excel_file = './data/workholic-data.xlsx'
    
    print("=== DEBUGGING LOGIN FLOW ===")
    
    # Check if file exists
    if not os.path.exists(excel_file):
        print("❌ Excel file does not exist")
        return
    
    print("✅ Excel file exists")
    
    try:
        # Load workbook
        workbook = openpyxl.load_workbook(excel_file)
        print(f"✅ Workbook loaded successfully")
        print(f"Sheet names: {workbook.sheetnames}")
        
        # Check if Employees sheet exists
        if 'Employees' not in workbook.sheetnames:
            print("❌ 'Employees' sheet not found")
            return
        
        print("✅ 'Employees' sheet found")
        
        # Read employees data
        employees_sheet = workbook['Employees']
        print(f"Max row: {employees_sheet.max_row}")
        print(f"Max column: {employees_sheet.max_column}")
        
        employees = []
        for row in employees_sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # If email exists
                employee = {
                    'email': row[0],
                    'name': row[1] or '',
                    'role': row[2],
                    'schedule': row[3],
                    'password': row[4] or None
                }
                employees.append(employee)
                print(f"Found employee: {employee['email']} - {employee['name']} - {employee['role']} - Password: {employee['password']}")
        
        print(f"\nTotal employees found: {len(employees)}")
        
        # Test specific login
        test_email = "admin@workholic.in"
        test_password = "admin123"
        
        print(f"\n=== TESTING LOGIN FOR {test_email} ===")
        
        employee = None
        for emp in employees:
            if emp['email'] == test_email:
                employee = emp
                break
        
        if not employee:
            print(f"❌ Employee {test_email} not found")
            return
        
        print(f"✅ Employee found: {employee['email']}")
        print(f"Password in database: '{employee['password']}'")
        print(f"Password to test: '{test_password}'")
        print(f"Passwords match: {employee['password'] == test_password}")
        
        if employee['password'] == test_password:
            print("✅ LOGIN SHOULD WORK!")
        else:
            print("❌ LOGIN WILL FAIL - Password mismatch")
            
    except Exception as e:
        print(f"❌ Error: {e}")

if __name__ == "__main__":
    debug_login_flow()
